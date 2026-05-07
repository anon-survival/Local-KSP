import torch
import util
import torch.nn.functional as F
import copy
from openpyxl import load_workbook
import time
import optim
import torch.nn as nn
import os
import random
import numpy as np

DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

def safe_logit(x, eps=1e-6):
    return torch.log((x + eps) / (1 - x + eps))

def postprocessing(args, cdf, is_dead, device='cpu', max_iters=10000, tol=1e-8, patience=100):
    EPS = 1e-8
    order = torch.argsort(cdf)
    cdf = cdf[order]
    cdf = cdf.unsqueeze(1)
    is_dead = is_dead[order].unsqueeze(1)
    N = cdf.shape[0]

    # Initialize learnable parameters
    a0_raw = torch.nn.Parameter(torch.tensor(0.0, device=device))
    b0 = torch.nn.Parameter(torch.tensor(0.0, device=device))
    alpha_raw = torch.nn.Parameter(torch.tensor(0.0, device=device))
    optimizer = torch.optim.Adam([a0_raw, b0, alpha_raw], lr=0.05)

    best_ks = float('inf')  # Track the best KS value
    best_params = None
    patience_counter = 0  # Initialize patience counter

    start_time = time.time()
    for iter in range(max_iters):

        with torch.set_grad_enabled(True):
            is_alive = (1 - is_dead).float()
            F_sorted = torch.sigmoid(torch.exp(a0_raw) * safe_logit(cdf) + b0) ** torch.exp(alpha_raw)

            denom = 1 - F_sorted + EPS
            weight = is_alive / denom
            F_weight = F_sorted * weight

            cum_weight = torch.cumsum(weight, dim=0)
            cum_F_weight = torch.cumsum(F_weight, dim=0)

            cum_weight_shifted = F.pad(cum_weight[:-1], (0, 0, 1, 0), value=0.0)
            cum_F_weight_shifted = F.pad(cum_F_weight[:-1], (0, 0, 1, 0), value=0.0)

            ecdf_cens = F_sorted * cum_weight_shifted - cum_F_weight_shifted
            ecdf_cens = torch.clamp(ecdf_cens, 0, N)

            ecdf_dead = torch.cumsum(is_dead, dim=0)
            ecdf_upper = (ecdf_dead + ecdf_cens) / N
            ecdf_upper = torch.clamp(ecdf_upper, 0, 1)
            ecdf_lower = ecdf_upper - is_dead / N

            KS_upper = torch.abs(ecdf_upper - F_sorted)
            KS_lower = torch.abs(ecdf_lower - F_sorted)
            KS_error = torch.max(torch.concat([KS_upper, KS_lower], dim=1), dim=1).values
            KS = torch.max(KS_error)

            print(f"KS: {KS.item():.5f}, Iteration: {iter+1}/{max_iters}", end="\r")

            optimizer.zero_grad()
            KS.backward()
            optimizer.step()

            # Gradient tolerance check
            grad_norm = torch.norm(torch.cat([a0_raw.grad.view(1), b0.grad.view(1), alpha_raw.grad.view(1)]))

            if grad_norm < tol:
                print(f"\nGradient norm below tolerance: {grad_norm:.6f}. Stopping early at iteration {iter+1}.")
                break

            # Early stopping check
            if torch.isfinite(KS):
                if KS.item() < best_ks:
                    best_ks = KS.item()
                    best_params = (a0_raw.clone(), b0.clone(), alpha_raw.clone())
                    patience_counter = 0
                else:
                    patience_counter += 1
                    if patience_counter >= patience:
                        print(f"\nEarly stopping at iteration {iter+1}. Best KS: {best_ks:.6f}")
                        break
            else:
                print("\nNon-finite KS detected. Restoring best and stopping.")
                break

    end_time = time.time()
    print("KSP time:", end_time - start_time)
    # workbook = load_workbook(filename='./ksp_time.xlsx')
    # sheet = workbook.active
    # last_row = sheet.max_row
    # sheet.cell(row=last_row+1, column=1, value=(end_time-start_time))
    # sheet.cell(row=last_row+1, column=2, value=(f'KSP_{args.dataset}_{args.model_dist}'))
    # workbook.save('./ksp_time.xlsx')

    # workbook = load_workbook(filename='./total_iter.xlsx')
    # sheet = workbook.active
    # last_row = sheet.max_row
    # sheet.cell(row=last_row+1, column=1, value=iter+1)
    # sheet.cell(row=last_row+1, column=2, value=(f'KSP_{args.dataset}_{args.model_dist}'))
    # workbook.save('./total_iter.xlsx')

    # Restore best parameters
    a0_raw, b0, alpha_raw = best_params
    a0 = torch.exp(a0_raw).item()
    b0 = b0.item()
    alpha = torch.exp(alpha_raw).item()
    print("Final parameters after early stopping:")
    print("a0:", a0)
    print("b0:", b0)
    print("alpha:", alpha)

    return a0, b0, alpha

def set_seed(seed=42):
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    np.random.seed(seed)
    random.seed(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

class ParamNet(nn.Module):
    def __init__(self, input_dim=1, hidden_dim=8):
        super().__init__()

        self.shared = nn.Sequential(
            nn.Linear(input_dim, hidden_dim),
            nn.ReLU(),
            nn.Linear(hidden_dim, hidden_dim),
            nn.ReLU()
        )

        self.a_head = nn.Linear(hidden_dim, 1)
        self.b_head = nn.Linear(hidden_dim, 1)
        self.alpha_head = nn.Linear(hidden_dim, 1)

        # self.a = nn.Parameter(torch.tensor(0.0))
        # self.b = nn.Parameter(torch.tensor(0.0))
        # self.alpha = nn.Parameter(torch.tensor(0.0))

    def forward(self, x):
        h = self.shared(x)

        a = F.softplus(self.a_head(h))
        b = self.b_head(h)
        alpha = F.softplus(self.alpha_head(h))

        # a = torch.exp(self.a)
        # b = self.b
        # alpha = torch.exp(self.alpha)

        return a, b, alpha
    
def kernel_postprocessing(args, cdf, cdf_matrix, is_dead, src, device='cpu', max_iters=10000, tol=1e-8, patience=200):
    set_seed(42)

    EPS = 1e-8
    order = torch.argsort(cdf)
    cdf = cdf[order].unsqueeze(1)
    is_dead = is_dead[order].unsqueeze(1)
    src = src[order]

    h_raw = torch.nn.Parameter(torch.tensor(0.0, device=device))
    # h_raw = torch.tensor(0.0, device=device)
    sigma_raw = torch.nn.Parameter(torch.tensor(0.0, device=device))

    model = ParamNet(input_dim=src.shape[1], hidden_dim=args.node).to(device)
        
    if args.use_kernel:
        # optimizer = optim.AdamWScheduleFreePaper(list(model.parameters()) + [sigma_raw], lr=0.01)
        optimizer = optim.AdamWScheduleFreePaper(list(model.parameters()) + [h_raw, sigma_raw], lr=0.03)
    else:
        optimizer = optim.AdamWScheduleFreePaper(list(model.parameters()) + [sigma_raw], lr=0.01)

    best_ks = float('inf')
    best_params = copy.deepcopy(model.state_dict())
    patience_counter = 0

    with torch.no_grad():
        import gower
        src_cpu = src.cpu().numpy()
        dij = torch.from_numpy(gower.gower_matrix(src_cpu)).to(device)
        alpha = args.n_quantile
        dij_no_diag = dij.clone()
        dij_no_diag.fill_diagonal_(float('inf'))
        di = torch.quantile(dij_no_diag, alpha, dim=1, keepdim=True)
        src_dist = dij / torch.sqrt(di * di.T + EPS)
        # src_dist = dij
        # h = torch.median(torch.triu(src_dist)[torch.triu(src_dist) != 0])

    start_time = time.time()
    optimizer.train()
    for iter in range(max_iters):
        with torch.set_grad_enabled(True):
            a, b, alpha = model(src.float())

            logit_cdf = safe_logit(cdf)
            if args.dataset == 'sequence' and args.model_dist == 'lognormal':
                logit_cdf = torch.clamp(logit_cdf, -7, 10)
            F_unsorted = torch.sigmoid(a * logit_cdf + b) ** alpha
            F_order = torch.argsort(F_unsorted.view(-1))
            F_sorted = F_unsorted.view(-1)[F_order].view(-1, 1)
            is_dead_sorted = is_dead.view(-1)[F_order].view(-1, 1)
            is_alive = (1 - is_dead_sorted).float()

            denom = 1 - F_sorted + EPS
            weight = is_alive / denom
            F_weight = F_sorted * weight

            if args.use_kernel:
                if args.sample:
                    g = torch.Generator(device=device)
                    g.manual_seed(42 + iter)
                    idx = torch.randperm(cdf.shape[0], generator=g, device=device)[:args.B]
                    src_dist_sorted = src_dist[F_order][:, F_order][:, idx]
                else:
                    src_dist_sorted = src_dist[F_order][:, F_order]

                if args.kernel == 'laplacian':
                    kernel_weight = torch.exp(-src_dist_sorted / torch.exp(h_raw)).to(device)
                    # kernel_weight = torch.exp(-src_dist_sorted / h).to(device)
                    kernel_weight = torch.clamp(kernel_weight, 0, 1)
                else:
                    kernel_weight = torch.exp(-src_dist_sorted / (2*torch.exp(h_raw)**2)).to(device)
                    kernel_weight = torch.clamp(kernel_weight, 0, 1)
                
                kw = kernel_weight * weight
                kf = kernel_weight * F_weight
                ksum = torch.sum(kernel_weight, dim=0)

                cum_weight = torch.cumsum(kw, dim=0)
                cum_F_weight = torch.cumsum(kf, dim=0)

                cum_weight_shifted = F.pad(cum_weight[:-1], (0, 0, 1, 0), value=0.0)
                cum_F_weight_shifted = F.pad(cum_F_weight[:-1], (0, 0, 1, 0), value=0.0)

                ecdf_cens = F_sorted * cum_weight_shifted - cum_F_weight_shifted
                ecdf_cens = torch.clamp(ecdf_cens, min=torch.zeros_like(ecdf_cens), max=ksum)

                ke = kernel_weight * is_dead_sorted
                ecdf_dead = torch.cumsum(ke, dim=0)

                ecdf_upper = (ecdf_dead + ecdf_cens) / ksum
                ecdf_upper = torch.clamp(ecdf_upper, 0, 1)

                ecdf_lower = ecdf_upper - ke / ksum

            else:
                cum_weight = torch.cumsum(weight, dim=0)
                cum_F_weight = torch.cumsum(F_weight, dim=0)

                cum_weight_shifted = F.pad(cum_weight[:-1], (0, 0, 1, 0), value=0.0)
                cum_F_weight_shifted = F.pad(cum_F_weight[:-1], (0, 0, 1, 0), value=0.0)

                ecdf_cens = F_sorted * cum_weight_shifted - cum_F_weight_shifted
                ecdf_cens = torch.clamp(ecdf_cens, min=0, max=F_sorted.shape[0])

                ecdf_dead = torch.cumsum(is_dead_sorted, dim=0)

                ecdf_upper = (ecdf_dead + ecdf_cens) / cdf.shape[0]
                ecdf_upper = torch.clamp(ecdf_upper, 0, 1)
                
                ecdf_lower = ecdf_upper - is_dead_sorted / cdf.shape[0]

            KS_upper = torch.abs(ecdf_upper - F_sorted)
            KS_lower = torch.abs(ecdf_lower - F_sorted)
            KS_error = torch.max(torch.max(KS_upper, KS_lower), dim=0).values

            # mono_penalty = rank_loss(args=args, cdf_before=cdf, cdf_after=F_unsorted, k=100, sigma_raw=sigma_raw, step=iter)
            idx = torch.argsort(cdf.view(-1))
            i = idx[:-1]
            j = idx[1:]
            diffs = F_unsorted.view(-1)[i] - F_unsorted.view(-1)[j]
            mono_penalty = F.relu(diffs).mean()

            # lambda_mono = max(0, 1 - iter / 10000)
            # lambda_mono = torch.exp(- torch.tensor([iter/100], device=device))
            # lambda_mono = 1/2 + 1/2 * np.cos(np.pi * (iter/400))
            KS = torch.mean(KS_error) + mono_penalty
            # if iter < 500:
            #     KS = torch.mean(KS_error) + mono_penalty
            # elif iter < 1000:
            #     KS = torch.mean(KS_error) + 0.5*mono_penalty
            # else:
            #     KS = torch.mean(KS_error)

            print(f"KS: {KS.item():.5f}, Iteration: {iter+1}/{max_iters}", end="\r")

            best_h_raw = h_raw.clone().detach()
            best_sigma_raw = sigma_raw.clone().detach()

            if torch.isfinite(KS):
                if KS.item() < best_ks:
                    best_ks = KS.item()
                    best_params = copy.deepcopy(model.state_dict())
                    best_h_raw = h_raw.clone().detach()
                    best_sigma_raw = sigma_raw.clone().detach()
                    patience_counter = 0
                else:
                    patience_counter += 1
                    if patience_counter >= patience:
                        print(f"\nEarly stopping at iteration {iter+1}. Best KS: {best_ks:.6f}")
                        break
            else:
                print("\nNon-finite KS detected. Restoring best and stopping.")
                break

            optimizer.zero_grad()
            KS.backward()

            all_params = list(model.parameters())
            if h_raw.requires_grad:
                all_params.append(h_raw)
            if sigma_raw.requires_grad:
                all_params.append(sigma_raw)

            grads = []
            for p in model.parameters():
                if p.grad is not None:
                    grads.append(p.grad.view(-1))
            if h_raw.grad is not None:
                grads.append(h_raw.grad.view(-1))
            if sigma_raw.grad is not None:
                grads.append(sigma_raw.grad.view(-1))

            if grads:
                grad_vec = torch.cat(grads)
                if not torch.isfinite(grad_vec).all():
                    print("\nNon-finite gradient detected. Skipping this step.")
                    optimizer.zero_grad()
                    continue
                grad_norm = torch.norm(grad_vec)
            else:
                grad_norm = torch.tensor(0.0, device=device)

            if grad_norm < tol:
                print(f"\nGradient norm below tolerance: {grad_norm:.6f}. Stopping early at iteration {iter+1}.")
                break
            
            optimizer.step()

    # optimizer.eval()
    
    end_time = time.time()
    # workbook = load_workbook(filename='./ksp_time.xlsx')
    # sheet = workbook.active
    # last_row = sheet.max_row
    # sheet.cell(row=last_row+1, column=1, value=(end_time-start_time))
    # sheet.cell(row=last_row+1, column=2, value=(f'kernel_KSP_{args.dataset}_{args.model_dist}'))
    # workbook.save('./ksp_time.xlsx')
    
    # workbook = load_workbook(filename='./total_iter.xlsx')
    # sheet = workbook.active
    # last_row = sheet.max_row
    # sheet.cell(row=last_row+1, column=1, value=iter+1)
    # sheet.cell(row=last_row+1, column=2, value=(f'kernel_KSP_{args.dataset}_{args.model_dist}'))
    # workbook.save('./total_iter.xlsx')

    print("---------------------------------------------------------------------")
    print("sigma:", torch.exp(best_sigma_raw).item())
    if args.use_kernel:
        print("h:", torch.exp(best_h_raw).item())

    # workbook = load_workbook(filename='./hyperparameter.xlsx')
    # sheet = workbook.active
    # last_row = sheet.max_row
    # sheet.cell(row=last_row+1, column=1, value=torch.exp(best_sigma_raw).item())
    # sheet.cell(row=last_row+1, column=2, value=torch.exp(best_h_raw).item())
    # sheet.cell(row=last_row+1, column=3, value=(f'kernel_KSP_{args.dataset}_{args.model_dist}'))
    # workbook.save('./hyperparameter.xlsx')

    print("---------------------------------------------------------------------")
    print("conditional KSP time:", end_time - start_time)
    print("---------------------------------------------------------------------")

    model.load_state_dict(best_params)

    return model


# def kernel_postprocessing(args, cdf, cdf_matrix, is_dead, src, device='cpu', max_iters=10000, tol=1e-8, patience=200):
#     set_seed(42)

#     EPS = 1e-8
#     cdf = cdf.unsqueeze(1)
#     is_dead = is_dead.unsqueeze(1)

#     model = ParamNet(input_dim=src.shape[1], hidden_dim=args.node).to(device)
#     sigma_raw = torch.nn.Parameter(torch.tensor(0.0, device=device))        

#     optimizer = optim.AdamWScheduleFreePaper(list(model.parameters()) + [sigma_raw], lr=0.01)

#     best_ks = float('inf')
#     best_params = copy.deepcopy(model.state_dict())
#     patience_counter = 0

#     start_time = time.time()
#     optimizer.train()
#     for iter in range(max_iters):
#         with torch.set_grad_enabled(True):
#             a, b, alpha = model(src.float())

#             logit_cdf = safe_logit(cdf)

#             if args.dataset == 'sequence' and args.model_dist == 'lognormal':
#                 logit_cdf = torch.clamp(logit_cdf, -7, 10)

#             F_unsorted = torch.sigmoid(a * logit_cdf + b) ** alpha
#             F_order = torch.argsort(F_unsorted.view(-1))
#             F_sorted = F_unsorted.view(-1)[F_order]
#             is_dead_sorted = is_dead.view(-1)[F_order]
#             is_alive = (1 - is_dead_sorted).float()

#             denom = 1 - F_sorted + EPS
#             weight = is_alive / denom
#             F_weight = F_sorted * weight
            
#             if args.use_kernel:
#                 cdf_weight = torch.sigmoid(a * safe_logit(cdf_matrix.float()) + b) ** alpha

#                 g = torch.Generator(device=device)
#                 g.manual_seed(42 + iter)
#                 idx = torch.randint(0, F_sorted.shape[0], (1,), generator=g, device=device)
#                 # kernel_weight = torch.norm(cdf_weight - cdf_weight[idx, :], args.n_quantile, dim=1) / F_sorted.shape[0]
#                 kernel_weight = torch.norm(cdf_weight - cdf_weight[idx, :], float('Inf'), dim=1) / F_sorted.shape[0]
#                 kernel_weight = torch.exp(-kernel_weight)
#                 kernel_weight = kernel_weight[F_order]

#                 kw = kernel_weight * weight
#                 kf = kernel_weight * F_weight

#                 ksum = torch.sum(kernel_weight)

#                 cum_weight = torch.cumsum(kw, dim=0)
#                 cum_F_weight = torch.cumsum(kf, dim=0)

#                 cum_weight_shifted = torch.cat([torch.zeros_like(cum_weight[:1]), cum_weight[:-1]], dim=0)
#                 cum_F_weight_shifted = torch.cat([torch.zeros_like(cum_F_weight[:1]), cum_F_weight[:-1]], dim=0)

#                 ecdf_cens = F_sorted * cum_weight_shifted - cum_F_weight_shifted
#                 ecdf_cens = torch.clamp(ecdf_cens, min=torch.zeros_like(ecdf_cens), max=ksum)

#                 ke = kernel_weight * is_dead_sorted
#                 ecdf_dead = torch.cumsum(ke, dim=0)

#                 ecdf_upper = (ecdf_dead + ecdf_cens) / ksum
#                 ecdf_upper = torch.clamp(ecdf_upper, 0, 1)

#                 ecdf_lower = ecdf_upper - ke / ksum

#             else:
#                 cum_weight = torch.cumsum(weight, dim=0)
#                 cum_F_weight = torch.cumsum(F_weight, dim=0)

#                 cum_weight_shifted = torch.cat([torch.zeros_like(cum_weight[:1]), cum_weight[:-1]], dim=0)
#                 cum_F_weight_shifted = torch.cat([torch.zeros_like(cum_F_weight[:1]), cum_F_weight[:-1]], dim=0)

#                 ecdf_cens = F_sorted * cum_weight_shifted - cum_F_weight_shifted
#                 ecdf_cens = torch.clamp(ecdf_cens, min=0, max=F_sorted.shape[0])

#                 ecdf_dead = torch.cumsum(is_dead_sorted, dim=0)

#                 ecdf_upper = (ecdf_dead + ecdf_cens) / cdf.shape[0]
#                 ecdf_upper = torch.clamp(ecdf_upper, 0, 1)
                
#                 ecdf_lower = ecdf_upper - is_dead_sorted / cdf.shape[0]

#             KS_upper = torch.abs(ecdf_upper - F_sorted)
#             KS_lower = torch.abs(ecdf_lower - F_sorted)
#             KS_error = torch.max(torch.maximum(KS_upper, KS_lower))

#             # diffs = F_unsorted - F_unsorted[idx].T
#             # correct_order = (cdf < cdf[idx].T)
#             # sigma = torch.exp(sigma_raw)
#             # if args.rank == 'softplus':
#             #     rank_loss = F.softplus(diffs/sigma)
#             # else:
#             #     rank_loss = torch.exp(diffs/sigma)
#             # mono_penalty = (rank_loss*correct_order).mean()

#             mono_penalty = rank_loss(args=args, cdf_before=cdf, cdf_after=F_unsorted, k=5, sigma_raw=sigma_raw, step=iter)

#             KS = KS_error
#             # KS = KS_error + mono_penalty

#             print(f"KS: {KS.item():.5f}, Iteration: {iter+1}/{max_iters}", end="\r")

#             best_sigma_raw = sigma_raw.clone().detach()

#             if torch.isfinite(KS):
#                 if KS.item() < best_ks:
#                     best_ks = KS.item()
#                     best_params = copy.deepcopy(model.state_dict())
#                     best_sigma_raw = sigma_raw.clone().detach()
#                     patience_counter = 0
#                 else:
#                     patience_counter += 1
#                     if patience_counter >= patience:
#                         print(f"\nEarly stopping at iteration {iter+1}. Best KS: {best_ks:.6f}")
#                         break
#             else:
#                 print("\nNon-finite KS detected. Restoring best and stopping.")
#                 break

#             optimizer.zero_grad()
#             KS.backward()

#             all_params = list(model.parameters())
#             if sigma_raw.requires_grad:
#                 all_params.append(sigma_raw)

#             grads = []
#             for p in model.parameters():
#                 if p.grad is not None:
#                     grads.append(p.grad.view(-1))
#             if sigma_raw.grad is not None:
#                 grads.append(sigma_raw.grad.view(-1))

#             if grads:
#                 grad_vec = torch.cat(grads)
#                 if not torch.isfinite(grad_vec).all():
#                     print("\nNon-finite gradient detected. Skipping this step.")
#                     optimizer.zero_grad()
#                     continue
#                 grad_norm = torch.norm(grad_vec)
#             else:
#                 grad_norm = torch.tensor(0.0, device=device)

#             if grad_norm < tol:
#                 print(f"\nGradient norm below tolerance: {grad_norm:.6f}. Stopping early at iteration {iter+1}.")
#                 break
            
#             optimizer.step()

#     # optimizer.eval()
    
#     end_time = time.time()
#     # workbook = load_workbook(filename='./ksp_time.xlsx')
#     # sheet = workbook.active
#     # last_row = sheet.max_row
#     # sheet.cell(row=last_row+1, column=1, value=(end_time-start_time))
#     # sheet.cell(row=last_row+1, column=2, value=(f'kernel_KSP_{args.dataset}_{args.model_dist}'))
#     # workbook.save('./ksp_time.xlsx')
    
#     # workbook = load_workbook(filename='./total_iter.xlsx')
#     # sheet = workbook.active
#     # last_row = sheet.max_row
#     # sheet.cell(row=last_row+1, column=1, value=iter+1)
#     # sheet.cell(row=last_row+1, column=2, value=(f'kernel_KSP_{args.dataset}_{args.model_dist}'))
#     # workbook.save('./total_iter.xlsx')

#     print("---------------------------------------------------------------------")
#     print("sigma:", torch.exp(best_sigma_raw).item())

#     # workbook = load_workbook(filename='./hyperparameter.xlsx')
#     # sheet = workbook.active
#     # last_row = sheet.max_row
#     # sheet.cell(row=last_row+1, column=1, value=torch.exp(best_sigma_raw).item())
#     # sheet.cell(row=last_row+1, column=2, value=torch.exp(best_h_raw).item())
#     # sheet.cell(row=last_row+1, column=3, value=(f'kernel_KSP_{args.dataset}_{args.model_dist}'))
#     # workbook.save('./hyperparameter.xlsx')

#     print("---------------------------------------------------------------------")
#     print("conditional KSP time:", end_time - start_time)
#     print("---------------------------------------------------------------------")

#     model.load_state_dict(best_params)

#     return model

# Random pair without replacement
def rank_loss(args, cdf_before, cdf_after, k, sigma_raw, step, base_seed=42):
    m = min(cdf_before.shape[0], k)

    g = torch.Generator(device=cdf_before.device)
    g.manual_seed(base_seed + step)

    # j_idx = torch.randperm(cdf_before.shape[0], device=cdf_before.device)[:m]
    k_idx = torch.randperm(cdf_before.shape[0], device=cdf_before.device, generator=g)[:m]

    # diffs = cdf_after[j_idx] - cdf_after[k_idx].T
    # correct_order = (cdf_before[j_idx] < cdf_before[k_idx].T)
    diffs = cdf_after - cdf_after[k_idx].T
    correct_order = (cdf_before < cdf_before[k_idx].T)

    sigma = torch.exp(sigma_raw)

    if args.rank == 'softplus':
        rank_loss = F.softplus(diffs/sigma)
    else:
        rank_loss = torch.exp(diffs/sigma)

    return (rank_loss*correct_order).mean()
