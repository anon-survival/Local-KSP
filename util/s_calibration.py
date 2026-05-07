import torch
from openpyxl import load_workbook

# def s_calibration(points, phase, is_dead, args, gamma=1.0, differentiable=False, device='cpu'):
#     new_is_dead = is_dead.detach().clone()
#     new_is_dead[points > 1. - 1e-4] = 1
#     points = points.to(device).view(-1, 1)

#     # NON-CENSORED POINTS
#     points_dead = points[new_is_dead.long() == 1]
#     # s = torch.tensor([0.05, 0.10, 0.20, 0.40, 0.50, 0.60, 0.80, 0.90, 0.95, 1.00], device=device)
#     # s = ((torch.arange(20) + 1) / 20).to(device)

#     if args.phase == 'test' or phase == 'valid':
#         s = ((torch.arange(20) + 1) / 20).to(device)

#     else:
#         s = torch.rand(args.num_s, device=device)
#         print(s)
#         # s = torch.distributions.Beta(args.alpha, args.beta).sample((args.num_s, )).to(device)
        
#     zeros = torch.zeros(s.shape[0]).to(device)
#     lower_diff_dead = points_dead - zeros
#     upper_diff_dead = s - points_dead
#     diff_product_dead = lower_diff_dead * upper_diff_dead
    
#     assert lower_diff_dead.shape == upper_diff_dead.shape, (lower_diff_dead.shape, upper_diff_dead.shape)
#     assert lower_diff_dead.shape == (points_dead.shape[0], s.shape[0])

#     if differentiable == True:
#         #soft_membership_dead = (points_dead <= s).float()
#         soft_membership_dead = torch.sigmoid(gamma * diff_product_dead)
        
#     else:
#         soft_membership_dead = (points_dead <= s).float()
        
#     fraction_dead = soft_membership_dead.sum(0)/points.shape[0]
    
#     # CENSORED POINTS
#     points_cens = points[new_is_dead.long() == 0]
#     upper_diff_for_soft_cens = s - points_cens
    
#     zeros = torch.zeros(s.shape[0]).to(device)
#     lower_diff_cens = points_cens - zeros
#     upper_diff_cens = s - points_cens
    
#     diff_product_cens = lower_diff_cens * upper_diff_cens
    
#     assert s.shape[0] == diff_product_cens.shape[1]
    
#     EPS = 1e-12
#     right_censored_interval_size = 1 - points_cens + EPS
    
#     if differentiable == True:
#         bin_index_one = torch.sigmoid(gamma * diff_product_cens)

#     else:
#         bin_index_one = (points_cens <= s).float()

#     upper_diff_within_bin = (upper_diff_for_soft_cens * bin_index_one)
#     partial_bin_assigned_weight = (upper_diff_within_bin/right_censored_interval_size).sum(0) / points.shape[0]
    
#     # if args.phase == 'test' and phase == 'test':
#     #     error = fraction_dead + partial_bin_assigned_weight

#     #     workbook = load_workbook(filename='./calibration_plot.xlsx')
#     #     sheet = workbook.active
#     #     last_row = sheet.max_row
#     #     for i in range(len(error)):
#     #         sheet.cell(row=last_row+1, column=i+1, value=error[i].item())
#     #     sheet.cell(row=last_row+1, column=21, value=f'{args.dataset}_{args.model_dist}')
#     #     workbook.save('./calibration_plot.xlsx')

#     return torch.pow(fraction_dead + partial_bin_assigned_weight - s, 2).sum() / s.shape[0]

def s_calibration(points, phase, is_dead, args, gamma=1.0, differentiable=False, device='cpu'):

    points = points.to(device).view(-1, 1)
    new_is_dead = is_dead.detach().clone()
    new_is_dead[points.view(-1) > 1. - 1e-4] = 1

    mask_dead = new_is_dead == 1
    mask_cens = ~mask_dead

    if args.phase == 'test' or phase == 'valid':
        s = ((torch.arange(20, device=device) + 1) / 20)
    else:
        s = torch.rand(args.num_s, device=device)
        # s = (1 - 0.05) * s + 0.05

    # ------------------------
    # DEAD
    # ------------------------
    points_dead = points[mask_dead]

    if points_dead.numel() > 0:
        diff_product_dead = points_dead * (s - points_dead)

        if differentiable:
            soft_dead = torch.sigmoid(gamma * diff_product_dead)
        else:
            soft_dead = (points_dead <= s).float()

        fraction_dead = soft_dead.sum(0) / points.shape[0]
    else:
        fraction_dead = torch.zeros_like(s)

    # ------------------------
    # CENSORED
    # ------------------------
    points_cens = points[mask_cens]

    if points_cens.numel() > 0:
        diff_product_cens = points_cens * (s - points_cens)

        if differentiable:
            bin_index = torch.sigmoid(gamma * diff_product_cens)
        else:
            bin_index = (points_cens <= s).float()

        EPS = 1e-12
        interval = 1 - points_cens + EPS

        upper = (s - points_cens) * bin_index
        partial_weight = (upper / interval).sum(0) / points.shape[0]
    else:
        partial_weight = torch.zeros_like(s)

    return ((fraction_dead + partial_weight - s) ** 2).mean()